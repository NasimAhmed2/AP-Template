import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime
import json
import os
from django.conf import settings
import sqlite3
from .models import InvoiceDetail
import shutil
from django.http import FileResponse, JsonResponse
import traceback


# Get today's date
today = datetime.now()
# Format the date as ddmmyyyy
formatted_date = today.strftime("%d%m%Y")

def ensure_files_and_directories(user_index):
    # Define base directory for user-specific files
    base_dir = os.path.join(settings.BASE_DIR, 'TemplateData', str(user_index))
    os.makedirs(base_dir, exist_ok=True)  # Ensure the directory exists

    # Define paths for dummy files
    dummy_template_path = os.path.join(settings.BASE_DIR, 'TemplateData', 'Templates.xlsx')
    dummy_header_path = os.path.join(settings.BASE_DIR, 'TemplateData', 'header.xlsx')

    # Define user-specific file paths
    user_template_path = os.path.join(base_dir, 'Templates.xlsx')
    user_header_path = os.path.join(base_dir, 'header.xlsx')

    # Check and copy Templates.xlsx
    if not os.path.exists(user_template_path):
        shutil.copy(dummy_template_path, user_template_path)
        print(f"Copied Templates.xlsx to: {user_template_path}")

    # Check and copy header.xlsx
    if not os.path.exists(user_header_path):
        shutil.copy(dummy_header_path, user_header_path)
        print(f"Copied header.xlsx to: {user_header_path}")

    # Return the paths
    return user_template_path, user_header_path

def template_formation(data,user_index,User):
    data_len = len(data)
    message = []
    if data_len > 0:
        
        file_path_template,file_path_header = ensure_files_and_directories(user_index)
        data_template = pd.read_excel(file_path_template, skiprows=1)
        # print(data_templateheader)
        Document_Number_list = data_template['BaseRef'].tolist()
        path = os.path.join(settings.BASE_DIR, 'GRN_Data', str(user_index), 'Open_GRN_Data.csv')
        # file_path_header = 'TemplateData/header.xlsx'
        
        # path = 'GRN_DATA/Open_GRN_Data.csv'
        data_grn = pd.read_csv(path)
        for invoice_name in data:
            file_name = f"{invoice_name}.json"
            response_dir = os.path.join(settings.MEDIA_ROOT, "responses", str(user_index))
            response_file = os.path.join(response_dir, file_name)
            try:
                # Check if the response file exists
                if os.path.exists(response_file):
                    # print(response_file)
                    # Load the response data from the file
                    with open(response_file, "r") as file:
                        api_response = json.load(file)
                        # print(api_response)
                    
                    # Extract relevant data
                    result = api_response.get("result", {})
                    # Extract default data
                    invoice_data = result.get('Invoice_data', {})
                    
                    # print('hello--1')
                    table_data_json = result.get('CHECKS', {}).get('table_data', {}).get('Table_Check_data', '[]')
                    table_data = json.loads(table_data_json) if isinstance(table_data_json, str) else table_data_json
                    table_data_df = pd.DataFrame(table_data)
                    # print('hello--2')
                    table_amount_sum = table_data_df['amount'].sum() or table_data_df['qty_unitprice'].sum()
                    total_amount_invoice = invoice_data.get('InvoiceTotal')
                    invoice_id = invoice_data.get('InvoiceId')
                    # print(invoice_id)
                    invoice_date = invoice_data.get('InvoiceDate')
                    po_number = invoice_data.get('PurchaseOrder')
                    # print('hello--3',po_number)
                    if po_number is not None:
                        po_number_ = re.sub(r'[^0-9]', '', po_number)
                        if len(po_number_) > 8:
                            po_number_ = po_number_[:8]
                    
                    # print('hello--4')
                    
                    if invoice_id:
                        # print(invoice_id)
                        # invoice_id = str(invoice_id).lstrip('0')
                        # print(data_grn['Supplier Ref No'])
                        invoice_id = str(invoice_id).lstrip('0').strip()  # Remove leading zeros & extra spaces
                        invoice_id_ = str(invoice_id).strip()  # Remove leading zeros & extra spaces
                        filtered_row = data_grn[
                                                data_grn['Supplier Ref No'].astype(str).str.strip().str.contains(invoice_id, case=False, na=False)
                                            ]

                        # print('this---1')
                        if filtered_row.empty:
                            print("No rows matched with invoice id")
                            message.append(f'No data found in open grn records against {invoice_id} for {invoice_name}')
                            
                        else:
                            
                            # print(filtered_row.iloc[0])
                            row_data = filtered_row.iloc[0].to_dict()  # Convert the row to a dictionary
                            DocNum = filtered_row.iloc[0]['Document Number'] 
                            if DocNum in Document_Number_list:
                                message.append(f"THIS {invoice_name} IS ALREADY IN TEMPLATE")
                                 
                            else:
                                # print('this----1')
                                DocNum = row_data.get('Document Number', ' ')
                                Series = row_data.get('Series', ' ')
                                CardCode = row_data.get('Customer/Supplier No.', ' ')

                                # Convert dates safely
                                date1 = row_data.get('Posting Date', ' ')
                                DocDate_formatted = pd.to_datetime(date1, errors='coerce').strftime("%Y%m%d") if date1 != ' ' else ' '

                                date2 = row_data.get('Due Date', ' ')
                                DocDueDate_formatted = pd.to_datetime(date2, errors='coerce').strftime("%Y%m%d") if date2 != ' ' else ' '

                                DocDate = DocDate_formatted
                                DocDueDate = DocDueDate_formatted

                                # Convert invoice_date safely
                                invoice_date_obj = datetime.strptime(invoice_date, "%Y-%m-%d") if 'invoice_date' in locals() else None
                                invoice_formatted_date = invoice_date_obj.strftime("%Y%m%d") if invoice_date_obj else ' '
                                TaxDate = invoice_formatted_date

                                DiscPrcnt = 0
                                DocCur = row_data.get('Currency Type Header') or row_data.get('Currency Type') or ' '
                                DocRate = ' '
                                NumAtCard = row_data.get('Supplier Ref No', ' ')
                                CntctCode = ' '
                                # print('this----5')
                                DocType = row_data.get('Document Type', ' ')
                                SlpCode = ' '
                                Comments = ' '
                                GSTTranType = row_data.get('GSTTransactionType', ' ')
                                # print('this----6')
                                new_data = [' ', Series,CardCode,DocDate,DocDueDate,TaxDate,DiscPrcnt,DocCur,
                                            DocRate,NumAtCard,CntctCode,DocType,SlpCode,Comments,GSTTranType]
                                # print('this----5')
                                # Load the workbook and select the active sheet
                                workbook = load_workbook(file_path_header)
                                sheet = workbook.active
                                # Append data to the next available row
                                sheet.append(new_data)

                                # Save the workbook
                                workbook.save(file_path_header)
                                print("Header Data appended successfully.")
                                # Resetting index and iterating
                                filtered_row_ = filtered_row.reset_index()
                                # print(filtered_row_)
                                for index, row in filtered_row_.iterrows():
                                    try:
                                        LineNum = index
                                        Due_Amount = row['Total Paymt Due']
                                        if DocCur == 'INR':
                                            Due_Amount = row['Total Paymt Due']
                                        else:
                                            Due_Amount = row['Total Payment Due FC']
                                        # print("hello--3",Due_Amount,total_amount_invoice)
                                        if abs(float(Due_Amount) - float(total_amount_invoice)) < 1 :
                                            # print("hello--4")
                                            # print(row)
                                            DocNum = row['Document Number']
                                            ItemCode = row['Item No.']
                                            Quantity = row['Quantity']
                                            Price = row['Price']
                                            TaxCode = row['Tax Code']
                                            BaseType = 20
                                            BaseEntry = row['GRPO DocEntry']
                                            # print(BaseEntry)
                                            if DocCur == 'INR':
                                                Price_ = row['Total Before Discount']
                                            else:
                                                Price_ = row['Total Before Discount FC']
                                            # Price_ = row['Total Before Discount']
                                            LocCode = 2
                                            WhsCode = row['WarehouseCode']
                                            CntctCode = ' '
                                            # print("hello--5")
                                            DocType = row['Document Type']
                                            SlpCode = ' '
                                            Comments = ' '
                                            GSTTranType = row['GSTTransactionType']
                                            new_data = [' ', LineNum,ItemCode,Quantity,Price,TaxCode,BaseType,BaseEntry,
                                                        LineNum,Price_,LocCode,WhsCode,int(DocNum)]
                                            
                                            # Load the workbook and select the active sheet
                                            workbook = load_workbook(file_path_template)
                                            sheet = workbook.active
                                            # Append data to the next available row
                                            sheet.append(new_data)

                                            # Save the workbook
                                            workbook.save(file_path_template)
                                            
                                            print(f"[✓] Templates appended successfully - Row: {index}, Invoice: {invoice_name}")
                                        else:
                                            print("Templates not appended succesfully")
                                    except:
                                        print({"error": str(e)})
                                        traceback.print_exc()
                                        print("Error:", e)

                                # # Connect to the SQLite database
                                # db_path = os.path.join(settings.BASE_DIR, "db.sqlite3")
                                # conn = sqlite3.connect(db_path)
                                # cursor = conn.cursor()

                                # # Update the okay_status and okay_message for the matching file_name
                                # update_query = """
                                # UPDATE invoice_detail
                                # SET status = ?
                                # WHERE file_name = ?
                                # """
                                # cursor.execute(update_query, ('confirmed', invoice_name))

                                # Commit changes to the database
                                # conn.commit()
                                try:
                                    # Get the invoice by file_name using Django's ORM
                                    invoice = InvoiceDetail.objects.get(user=User, file_name=invoice_name)

                                    # Update the fields
                                    invoice.status = 'confirmed'
                                    

                                    # Save the changes
                                    invoice.save()
                                    

                                except InvoiceDetail.DoesNotExist:
                                    # If the invoice does not exist
                                    response = {
                                        "message": f"No invoice found with the name '{invoice_name}'",
                                        "status": "error"
                                    }
                                    return JsonResponse(response)

                                
                                message.append(f"Templates appended succesfully for {invoice_name}")
                    
                    elif po_number_:
                        pass
                    else:
                        pass
            except Exception as e:
                print({"error": str(e)})
                traceback.print_exc()
                print("Error:", e)

            print(file_name)
    return message

    # path_ = 'TemplateData/header.xlsx'
    # data_templateheader = pd.read_excel(path_, skiprows=1)
    # doc_number_list = data_templateheader['DocNum'].tolist()
    # path = 'GRN_DATA/Open_GRN_Data.csv'
    # data_grn = pd.read_csv(path)

    # invoice_data = data['invoice_data']
    # table_data = data['table_data']['tabledata_']
    # table_data_df = pd.DataFrame(table_data)
    # amount_list = table_data_df['amount'].tolist() or table_data_df['qty_unitprice'].tolist()
    # po_number = invoice_data.get('PurchaseOrder')
    # invoice_date = invoice_data.get('InvoiceDate')
    # # formatted_date_invoice = invoice_date.strftime("%d%m%Y")
    # invoice_currency = invoice_data.get('Currency')
    # po_number_ = re.sub(r'[^0-9]', '', po_number)
    # if len(po_number_) > 8:
    #     po_number_ = po_number_[:8]
    # print(po_number_)
    # if po_number_:
    #     filtered_row = data_grn[data_grn['Remarks'].str.contains(po_number_, na=False)]
    #     # print(filtered_row)
    #     if filtered_row.empty:
    #         message = 'No data found in open grn records'
    #         return message
    #     else:
    #         file_path_header = 'TemplateData/header.xlsx'
    #         file_path_template = 'TemplateData/Templates.xlsx'
    #         # print(filtered_row.iloc[0])
    #         DocNum = filtered_row.iloc[0]['Document Number'] 
    #         if DocNum in doc_number_list:
    #             return "THIS INVOICE IS ALREADY IN TEMPLATE"
    #         else:
    #             DocNum = filtered_row.iloc[0]['Document Number']
    #             Series = filtered_row.iloc[0]['Series']
    #             CardCode = filtered_row.iloc[0]['Customer/Supplier No.']
    #             DocDate = formatted_date
    #             DocDueDate = formatted_date
    #             TaxDate = formatted_date
    #             DiscPrcnt = 0
    #             DocCur = filtered_row.iloc[0]['Currency Type']
    #             DocRate = ' '
    #             NumAtCard = filtered_row.iloc[0]['Supplier Ref No']
    #             CntctCode = ' '
                
    #             DocType = filtered_row.iloc[0]['Document Type']
    #             SlpCode = ' '
    #             Comments = ' '
    #             GSTTranType = filtered_row.iloc[0]['GSTTransactionType']
    #             new_data = [int(DocNum), int(Series),CardCode,DocDate,DocDueDate,TaxDate,DiscPrcnt,DocCur,
    #                         DocRate,NumAtCard,CntctCode,DocType,SlpCode,Comments,GSTTranType]
                
    #             # Load the workbook and select the active sheet
    #             workbook = load_workbook(file_path_header)
    #             sheet = workbook.active
    #             # Append data to the next available row
    #             sheet.append(new_data)

    #             # Save the workbook
    #             workbook.save(file_path_header)
    #             print("Data appended successfully.")
    #             LineNum = 0
    #             for amount in amount_list:
    #                 row = filtered_row[filtered_row['Total Before Discount'] == float(amount)]
    #                 print(row)
    #                 if row.empty:
    #                     pass
    #                 else:
    #                     DocNum = row['Document Number'].iloc[0]
                        
    #                     ItemCode = row['Item No.'].iloc[0]
    #                     Quantity = row['Quantity'].iloc[0]
    #                     Price = row['Price'].iloc[0]
    #                     TaxCode = row['Tax Code'].iloc[0]
    #                     BaseType = 20
    #                     BaseEntry = 8112
    #                     Price_ = row['Total Before Discount'].iloc[0]
    #                     LocCode = ' '
    #                     WhsCode = row['WarehouseCode'].iloc[0]
    #                     CntctCode = ' '
                        
    #                     DocType = row['Document Type'].iloc[0]
    #                     SlpCode = ' '
    #                     Comments = ' '
    #                     GSTTranType = row['GSTTransactionType'].iloc[0]
    #                     new_data = [int(DocNum), LineNum,ItemCode,Quantity,Price,TaxCode,BaseType,BaseEntry,
    #                                 LineNum,Price_,LocCode,WhsCode,int(DocNum)]
                        
    #                     # Load the workbook and select the active sheet
    #                     workbook = load_workbook(file_path_template)
    #                     sheet = workbook.active
    #                     # Append data to the next available row
    #                     sheet.append(new_data)

    #                     # Save the workbook
    #                     workbook.save(file_path_template)
    #                     LineNum = LineNum + 1
    #             return "Templates appended succesfully"
                    
    # return 'No po number foubnd on onvoice to map with grn data'

def retain_two_rows(path):
    # Paths to the two Excel files
    # file1_path = "TemplateData/header.xlsx"
    # file2_path = "TemplateData/Templates.xlsx"
    for file in path:
        try:
            # Load the workbook
            wb = load_workbook(file)

            # Select the active sheet or a specified sheet
            sheet = wb.active

            # Get the maximum row in the sheet
            max_row = sheet.max_row

            # Delete rows after the first two (header and first data row)
            if max_row > 2:
                sheet.delete_rows(3, max_row - 2)

            # Save the workbook back to the same file
            wb.save(file)
            print('Sheet updated')

        except Exception as e:
            print(f"Error: {str(e)}")